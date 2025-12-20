"""文件处理服务"""
import os
import uuid
import mimetypes
from datetime import datetime
from typing import List, Dict, Optional
from ..database import get_session
from ..models import UploadedFile
from ..config import Config


class FileExtractor:
    """文件文本提取器"""
    
    # 支持的文件扩展名及其处理方式
    SUPPORTED_EXTENSIONS = {
        # 纯文本文件
        '.txt': 'text',
        '.md': 'text',
        '.py': 'text',
        '.json': 'text',
        '.js': 'text',
        '.ts': 'text',
        '.html': 'text',
        '.css': 'text',
        '.xml': 'text',
        '.yaml': 'text',
        '.yml': 'text',
        '.ini': 'text',
        '.conf': 'text',
        '.cfg': 'text',
        '.log': 'text',
        '.csv': 'text',
        '.sql': 'text',
        '.sh': 'text',
        '.bat': 'text',
        '.java': 'text',
        '.c': 'text',
        '.cpp': 'text',
        '.h': 'text',
        '.go': 'text',
        '.rs': 'text',
        '.rb': 'text',
        '.php': 'text',
        # 文档文件
        '.pdf': 'pdf',
        '.docx': 'docx',
        '.xlsx': 'xlsx',
    }
    
    # DeepSeek 上下文限制 (128K tokens ≈ 约 400K 字符，保守估计 350K)
    MAX_TEXT_LENGTH = 350000
    
    def __init__(self):
        pass
    
    def is_supported(self, extension: str) -> bool:
        """检查文件扩展名是否支持"""
        return extension.lower() in self.SUPPORTED_EXTENSIONS
    
    def get_supported_extensions(self) -> list:
        """获取支持的文件扩展名列表"""
        return list(self.SUPPORTED_EXTENSIONS.keys())
    
    def extract(self, file_path: str, extension: str) -> tuple[str, str]:
        """
        根据文件扩展名提取文本
        
        Args:
            file_path: 文件路径
            extension: 文件扩展名
            
        Returns:
            tuple: (提取的文本, 状态: 'success'/'failed'/'too_large')
        """
        handler = self.SUPPORTED_EXTENSIONS.get(extension.lower())
        
        if not handler:
            return '', 'failed'
        
        try:
            if handler == 'text':
                text = self._extract_text(file_path)
            elif handler == 'pdf':
                text = self._extract_pdf(file_path)
            elif handler == 'docx':
                text = self._extract_docx(file_path)
            elif handler == 'xlsx':
                text = self._extract_xlsx(file_path)
            else:
                return '', 'failed'
            
            # 检查文本长度
            if len(text) > self.MAX_TEXT_LENGTH:
                return text[:1000] + '\n\n... [文件内容过长，已截断] ...', 'too_large'
            
            return text, 'success'
            
        except Exception as e:
            print(f"Extract file error: {e}")
            return '', 'failed'
    
    def _extract_text(self, file_path: str) -> str:
        """提取纯文本文件"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
            except Exception as e:
                raise e
        
        # 最后尝试二进制读取并忽略错误
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    
    def _extract_pdf(self, file_path: str) -> str:
        """提取PDF文本"""
        import pdfplumber
        
        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text:
                    text_parts.append(f"--- 第 {i+1} 页 ---\n{text}")
        
        return '\n\n'.join(text_parts)
    
    def _extract_docx(self, file_path: str) -> str:
        """提取Word文档文本"""
        from docx import Document
        
        doc = Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return '\n\n'.join(paragraphs)
    
    def _extract_xlsx(self, file_path: str) -> str:
        """提取Excel文件文本"""
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = [f"=== 工作表: {sheet_name} ==="]
            
            for row in sheet.iter_rows(values_only=True):
                # 过滤空值并转换为字符串
                row_values = [str(cell) if cell is not None else '' for cell in row]
                if any(row_values):  # 至少有一个非空值
                    sheet_text.append(' | '.join(row_values))
            
            if len(sheet_text) > 1:  # 有实际内容
                text_parts.append('\n'.join(sheet_text))
        
        wb.close()
        return '\n\n'.join(text_parts)


class FileService:
    """文件服务"""
    
    def __init__(self):
        self.config = Config()
        self.extractor = FileExtractor()
        
        # 文件上传目录
        self.upload_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            'uploads'
        )
        
        # 确保上传目录存在
        if not os.path.exists(self.upload_dir):
            os.makedirs(self.upload_dir)
    
    def get_user_upload_dir(self, user_id: int) -> str:
        """获取用户的上传目录"""
        user_dir = os.path.join(self.upload_dir, f'user_{user_id}')
        if not os.path.exists(user_dir):
            os.makedirs(user_dir)
        return user_dir
    
    def validate_file(self, filename: str, file_size: int) -> tuple[bool, str]:
        """
        验证文件
        
        Args:
            filename: 文件名
            file_size: 文件大小(字节)
            
        Returns:
            tuple: (是否有效, 错误信息)
        """
        if not filename:
            return False, '文件名不能为空'
        
        # 获取扩展名
        _, ext = os.path.splitext(filename)
        ext = ext.lower()
        
        # 检查扩展名
        if not self.extractor.is_supported(ext):
            supported = ', '.join(self.extractor.get_supported_extensions())
            return False, f'不支持的文件类型，支持: {supported}'
        
        # 检查文件大小 (100MB)
        max_size = 100 * 1024 * 1024
        if file_size > max_size:
            return False, f'文件过大，最大支持 100MB'
        
        return True, ''
    
    def save_file(self, user_id: int, file) -> dict:
        """
        保存上传的文件
        
        Args:
            user_id: 用户ID
            file: Flask 文件对象
            
        Returns:
            dict: {'success': bool, 'file_id': int, 'error': str}
        """
        try:
            filename = file.filename
            
            # 获取文件大小
            file.seek(0, 2)  # 移动到文件末尾
            file_size = file.tell()
            file.seek(0)  # 重置到文件开头
            
            # 验证文件
            is_valid, error = self.validate_file(filename, file_size)
            if not is_valid:
                return {'success': False, 'error': error}
            
            # 生成存储文件名
            _, ext = os.path.splitext(filename)
            ext = ext.lower()
            stored_filename = f'{uuid.uuid4().hex}{ext}'
            
            # 获取用户上传目录
            user_dir = self.get_user_upload_dir(user_id)
            file_path = os.path.join(user_dir, stored_filename)
            
            # 保存文件
            file.save(file_path)
            
            # 获取MIME类型
            mime_type, _ = mimetypes.guess_type(filename)
            if not mime_type:
                mime_type = 'application/octet-stream'
            
            # 提取文本
            extracted_text, extraction_status = self.extractor.extract(file_path, ext)
            text_length = len(extracted_text) if extracted_text else 0
            
            # 保存到数据库
            db = get_session()
            try:
                uploaded_file = UploadedFile(
                    user_id=user_id,
                    original_filename=filename,
                    stored_filename=stored_filename,
                    file_path=file_path,
                    file_size=file_size,
                    file_type=mime_type,
                    file_extension=ext,
                    extracted_text=extracted_text,
                    text_length=text_length,
                    extraction_status=extraction_status,
                    error_message=None if extraction_status == 'success' else '文本提取失败' if extraction_status == 'failed' else '文件内容过长'
                )
                db.add(uploaded_file)
                db.commit()
                db.refresh(uploaded_file)
                
                return {
                    'success': True,
                    'file_id': uploaded_file.id,
                    'filename': filename,
                    'file_size': file_size,
                    'extraction_status': extraction_status,
                    'text_length': text_length
                }
            except Exception as e:
                db.rollback()
                # 删除已保存的文件
                if os.path.exists(file_path):
                    os.remove(file_path)
                raise e
            finally:
                db.close()
                
        except Exception as e:
            print(f"Save file error: {e}")
            return {'success': False, 'error': f'保存文件失败: {str(e)}'}
    
    def get_file(self, file_id: int, user_id: int) -> dict:
        """
        获取文件信息
        
        Args:
            file_id: 文件ID
            user_id: 用户ID
            
        Returns:
            dict: 文件信息
        """
        db = get_session()
        try:
            file_obj = db.query(UploadedFile).filter(
                UploadedFile.id == file_id,
                UploadedFile.user_id == user_id
            ).first()
            
            if not file_obj:
                return None
            
            return {
                'id': file_obj.id,
                'original_filename': file_obj.original_filename,
                'file_size': file_obj.file_size,
                'file_type': file_obj.file_type,
                'file_extension': file_obj.file_extension,
                'text_length': file_obj.text_length,
                'extraction_status': file_obj.extraction_status,
                'created_at': file_obj.created_at.isoformat() if file_obj.created_at else None
            }
        finally:
            db.close()
    
    def get_file_text(self, file_id: int, user_id: int) -> str:
        """
        获取文件的提取文本
        
        Args:
            file_id: 文件ID
            user_id: 用户ID
            
        Returns:
            str: 提取的文本，如果文件不存在或无权限返回 None
        """
        db = get_session()
        try:
            file_obj = db.query(UploadedFile).filter(
                UploadedFile.id == file_id,
                UploadedFile.user_id == user_id
            ).first()
            
            if not file_obj:
                return None
            
            return file_obj.extracted_text
        finally:
            db.close()
    
    def get_user_files(self, user_id: int, limit: int = 50) -> list:
        """
        获取用户的文件列表
        
        Args:
            user_id: 用户ID
            limit: 返回数量限制
            
        Returns:
            list: 文件列表
        """
        db = get_session()
        try:
            files = db.query(UploadedFile).filter(
                UploadedFile.user_id == user_id
            ).order_by(
                UploadedFile.created_at.desc()
            ).limit(limit).all()
            
            return [{
                'id': f.id,
                'original_filename': f.original_filename,
                'file_size': f.file_size,
                'file_extension': f.file_extension,
                'text_length': f.text_length,
                'extraction_status': f.extraction_status,
                'created_at': f.created_at.isoformat() if f.created_at else None
            } for f in files]
        finally:
            db.close()
    
    def delete_file(self, file_id: int, user_id: int) -> dict:
        """
        删除文件
        
        Args:
            file_id: 文件ID
            user_id: 用户ID
            
        Returns:
            dict: {'success': bool, 'error': str}
        """
        db = get_session()
        try:
            file_obj = db.query(UploadedFile).filter(
                UploadedFile.id == file_id,
                UploadedFile.user_id == user_id
            ).first()
            
            if not file_obj:
                return {'success': False, 'error': '文件不存在或无权限'}
            
            # 删除物理文件
            if os.path.exists(file_obj.file_path):
                os.remove(file_obj.file_path)
            
            # 删除数据库记录
            db.delete(file_obj)
            db.commit()
            
            return {'success': True}
        except Exception as e:
            db.rollback()
            print(f"Delete file error: {e}")
            return {'success': False, 'error': f'删除文件失败: {str(e)}'}
        finally:
            db.close()
    
    def format_file_context(self, file_id: int, user_id: int) -> Optional[str]:
        """
        格式化文件内容为上下文
        
        Args:
            file_id: 文件ID
            user_id: 用户ID
            
        Returns:
            str: 格式化的文件上下文，如果文件不存在或提取失败则返回None
        """
        db = get_session()
        try:
            file_obj = db.query(UploadedFile).filter(
                UploadedFile.id == file_id,
                UploadedFile.user_id == user_id
            ).first()
            
            if not file_obj:
                return None
            
            if file_obj.extraction_status == 'failed':
                return f"[文件: {file_obj.original_filename}]\n(文本提取失败，无法读取文件内容)"
            
            if file_obj.extraction_status == 'too_large':
                return f"[文件: {file_obj.original_filename}]\n(文件内容过长，以下为部分内容)\n\n{file_obj.extracted_text}"
            
            return f"[文件: {file_obj.original_filename}]\n\n{file_obj.extracted_text}"
        finally:
            db.close()
    
    def get_file_contexts_from_ids(self, file_ids: List[int], user_id: int) -> str:
        """
        从文件ID列表获取格式化的文件上下文
        
        Args:
            file_ids: 文件ID列表
            user_id: 用户ID
            
        Returns:
            格式化的文件上下文字符串
        """
        if not file_ids:
            return ""
        
        contexts = []
        for file_id in file_ids:
            context = self.format_file_context(file_id, user_id)
            if context:
                contexts.append(context)
        
        if not contexts:
            return ""
        
        return "\n\n【参考文件】\n请仔细阅读以下文件内容，然后回答问题。\n\n" + "\n\n---\n\n".join(contexts) + "\n\n【问题】\n\n"
    
    def enrich_history_messages_with_files(self, history_messages: List[Dict], user_id: int) -> List[Dict]:
        """
        丰富历史消息，将文件上下文添加到对应的消息中
        
        Args:
            history_messages: 历史消息列表，格式为 [{'role': 'user', 'content': '...', 'files': [...]}, ...]
                            注意：传入的消息列表已经由 LangChain Memory 进行了窗口限制，无需再次检查轮数
            user_id: 用户ID
            
        Returns:
            丰富后的历史消息列表
        """
        # 收集历史文件上下文
        history_file_contexts = {}
        
        # 从后往前遍历
        for idx in range(len(history_messages) - 1, -1, -1):
            msg = history_messages[idx]
            
            # 只处理用户消息
            if msg.get('role') != 'user':
                continue
            
            # 检查是否有文件
            files = msg.get('files', [])
            if not files:
                continue
            
            # 收集文件ID
            file_ids = []
            for file_info in files:
                file_id = file_info.get('id')
                if file_id:
                    file_ids.append(file_id)
            
            if file_ids:
                context = self.get_file_contexts_from_ids(file_ids, user_id)
                if context:
                    history_file_contexts[idx] = context
        
        # 丰富消息
        enriched_messages = []
        for idx, msg in enumerate(history_messages):
            enriched_msg = msg.copy()
            
            # 如果是用户消息且有文件上下文，添加到content中
            if idx in history_file_contexts:
                file_context = history_file_contexts[idx]
                # 将文件上下文添加到消息内容的前面
                enriched_msg['content'] = file_context + enriched_msg.get('content', '')
            
            enriched_messages.append(enriched_msg)
        
        return enriched_messages

